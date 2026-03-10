"""Parse TCBS 'Lich su giao dich co phieu' XLSX exports."""

from typing import BinaryIO

import openpyxl
import pandas as pd

COLUMN_MAP = {
    "Mã CP": "ticker",
    "Ngày GD": "trading_date",
    "Giao dịch": "trade_side",
    "KL đặt": "order_volume",
    "Giá đặt": "order_price",
    "KL khớp": "matched_volume",
    "Giá khớp": "matched_price",
    "Giá trị khớp": "matched_value",
    "Phí": "fee",
    "Thuế": "tax",
    "Giá vốn": "cost_basis",
    "Lãi lỗ": "return_pnl",
    "Kênh GD": "channel",
    "Trạng thái": "status",
    "Loại lệnh": "order_type",
    "Số hiệu lệnh": "order_no",
}

FOOTER_MARKERS = {"Tổng", "Ngày xuất báo cáo", "Report Date", "Chú thích", "Total"}


def _extract_vn_prefix(cell_text: str) -> str:
    """Extract the Vietnamese prefix from a bilingual header cell.

    TCBS exports may use bilingual headers like 'Mã CP\\nTicker'.
    This returns just the part before the newline, stripped.
    Also strips any parenthetical suffixes like '(VND)'.
    """
    prefix = cell_text.split("\n")[0].strip()
    # Remove trailing parenthetical like " (VND)" or " (vnd)"
    if "(" in prefix:
        prefix = prefix[:prefix.index("(")].strip()
    return prefix


def extract_account_type(file: BinaryIO) -> str:
    """Read the account type from TCBS header rows.

    Scans first 15 rows for account type keywords.
    Returns 'margin', 'normal', or 'unknown'.
    """
    file.seek(0)
    workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
    worksheet = workbook.active

    for _row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=15, values_only=True)):
        for cell in row:
            cell_text = str(cell) if cell else ""
            if "Ký quỹ" in cell_text or "Ky quy" in cell_text:
                workbook.close()
                file.seek(0)
                return "margin"
            if "Thường" in cell_text or "Thuong" in cell_text:
                workbook.close()
                file.seek(0)
                return "normal"

    workbook.close()
    file.seek(0)
    return "unknown"


def _detect_header_row(file: BinaryIO) -> int:
    """Auto-detect the header row by scanning for known column names.

    Handles bilingual headers like 'Mã CP\\nTicker' by extracting the
    Vietnamese prefix before the newline. Matches against COLUMN_MAP keys.
    Returns the 0-indexed row number, defaults to 14 if not found.
    """
    file.seek(0)
    workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
    worksheet = workbook.active

    known_headers = set(COLUMN_MAP.keys())

    for row_index, row in enumerate(worksheet.iter_rows(
        min_row=1, max_row=30, values_only=True
    )):
        prefixes = set()
        for cell in row:
            if cell is not None:
                prefix = _extract_vn_prefix(str(cell))
                prefixes.add(prefix)
        matches = prefixes & known_headers
        if len(matches) >= 3:
            workbook.close()
            file.seek(0)
            return row_index

    workbook.close()
    file.seek(0)
    return 14  # fallback to original default


def _build_column_rename_map(actual_columns: list) -> dict:
    """Build a rename map from actual DataFrame columns to English names.

    Handles both exact matches ('Mã CP') and bilingual headers
    ('Mã CP\\nTicker') by extracting the Vietnamese prefix.
    """
    rename_map = {}
    for actual_col in actual_columns:
        actual_str = str(actual_col)
        # Try exact match first
        if actual_str in COLUMN_MAP:
            rename_map[actual_col] = COLUMN_MAP[actual_str]
            continue
        # Try prefix match (bilingual headers)
        prefix = _extract_vn_prefix(actual_str)
        if prefix in COLUMN_MAP:
            rename_map[actual_col] = COLUMN_MAP[prefix]
    return rename_map


def _find_order_no_column(actual_columns: list) -> dict:
    """Find the actual column name for order_no to set string dtype.

    Returns a dtype dict for pd.read_excel.
    """
    for col in actual_columns:
        col_str = str(col)
        prefix = _extract_vn_prefix(col_str)
        if prefix == "Số hiệu lệnh" or col_str == "Số hiệu lệnh":
            return {col_str: str}
    return {"Số hiệu lệnh": str}


def parse_tcbs_xlsx(file: BinaryIO) -> pd.DataFrame:
    """Parse a TCBS trade history XLSX into a normalized DataFrame.

    - Auto-detects header row by scanning for Vietnamese column names
    - Handles bilingual headers (e.g. 'Mã CP\\nTicker')
    - Maps column names to English
    - Filters footer/total rows
    - Ensures order_no is string type
    """
    file.seek(0)

    header_row = _detect_header_row(file)

    # First pass: read just header to find the order_no column name
    file.seek(0)
    header_df = pd.read_excel(file, header=header_row, nrows=0, engine="openpyxl")
    dtype_map = _find_order_no_column(list(header_df.columns))

    # Second pass: read full data with correct dtype
    file.seek(0)
    dataframe = pd.read_excel(
        file,
        header=header_row,
        dtype=dtype_map,
        engine="openpyxl",
    )

    # Rename columns (handles bilingual headers)
    rename_map = _build_column_rename_map(list(dataframe.columns))
    dataframe = dataframe.rename(columns=rename_map)

    # Keep only mapped columns that exist
    valid_columns = [value for value in COLUMN_MAP.values() if value in dataframe.columns]
    dataframe = dataframe[valid_columns]

    # Filter footer rows: remove NaN tickers and rows matching footer markers
    if "ticker" in dataframe.columns:
        dataframe = dataframe[dataframe["ticker"].notna()]
        dataframe = dataframe[
            ~dataframe["ticker"].astype(str).str.strip().isin(FOOTER_MARKERS)
        ]
        dataframe = dataframe[
            ~dataframe["ticker"]
            .astype(str)
            .str.contains("Ngày xuất|Report Date|Chú thích", na=False)
        ]

    # Ensure order_no is string
    if "order_no" in dataframe.columns:
        dataframe["order_no"] = dataframe["order_no"].astype(str).str.strip()

    dataframe = dataframe.reset_index(drop=True)
    return dataframe
